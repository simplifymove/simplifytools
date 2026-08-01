import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from PyPDF2 import PdfReader, PdfWriter

from python.engines.pdf_core import PdfCoreEngine
from python.engines.spreadsheet_engine import SpreadsheetConvertEngine
from python.engines.structured_data_engine import StructuredDataEngine


class ExcelToCsvTests(unittest.TestCase):
    def _workbook(self, path: Path):
        workbook = openpyxl.Workbook()
        first = workbook.active
        first.title = 'A B'
        first.append(['sku', 'quantity'])
        first.append(['001', 4])

        duplicate_after_sanitizing = workbook.create_sheet('A_B')
        duplicate_after_sanitizing.append(['name'])
        duplicate_after_sanitizing.append(['second'])

        empty = workbook.create_sheet('日本')
        workbook.save(path)

    def test_first_sheet_still_exports_one_csv(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / 'input.xlsx'
            output = Path(directory) / 'output.csv'
            self._workbook(source)

            SpreadsheetConvertEngine().convert(
                'excel-to-csv', str(source), str(output),
                {'sheet_mode': 'first', 'delimiter': 'comma'},
            )

            text = output.read_text(encoding='utf-8-sig')
            self.assertIn('sku,quantity', text)
            self.assertIn('001,4', text)
            self.assertFalse(zipfile.is_zipfile(output))

    def test_all_sheets_are_safe_unique_csv_members_in_zip(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / 'input.xlsx'
            output = Path(directory) / 'output.zip'
            self._workbook(source)

            SpreadsheetConvertEngine().convert(
                'excel-to-csv', str(source), str(output),
                {'sheet_mode': 'zip', 'delimiter': 'semicolon'},
            )

            self.assertTrue(zipfile.is_zipfile(output))
            with zipfile.ZipFile(output) as archive:
                self.assertEqual(archive.namelist(), ['A_B.csv', 'A_B_2.csv', 'sheet_3.csv'])
                self.assertIn('sku;quantity', archive.read('A_B.csv').decode('utf-8-sig'))
                self.assertEqual(archive.read('sheet_3.csv').decode('utf-8-sig'), '\n')


class JsonToXmlTests(unittest.TestCase):
    def _convert(self, directory: str, options, source_text='[{"id": 1}, {"id": 2}]'):
        source = Path(directory) / 'input.json'
        output = Path(directory) / 'output.xml'
        source.write_text(source_text, encoding='utf-8')
        StructuredDataEngine().convert('json-to-xml', str(source), str(output), options)
        return output

    def test_default_and_custom_root_item_tags(self):
        with tempfile.TemporaryDirectory() as directory:
            default_output = self._convert(directory, {})
            default_root = ET.parse(default_output).getroot()
            self.assertEqual(default_root.tag, 'root')
            self.assertEqual([child.tag for child in default_root], ['item', 'item'])

            custom_output = self._convert(directory, {'rootTag': 'inventory', 'itemTag': 'product'})
            custom_root = ET.parse(custom_output).getroot()
            self.assertEqual(custom_root.tag, 'inventory')
            self.assertEqual([child.tag for child in custom_root], ['product', 'product'])
            self.assertEqual([child.findtext('id') for child in custom_root], ['1', '2'])

    def test_empty_tags_use_defaults_and_invalid_tags_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            output = self._convert(directory, {'rootTag': '', 'itemTag': ''})
            self.assertEqual(ET.parse(output).getroot().tag, 'root')

            with self.assertRaisesRegex(ValueError, 'valid XML element name'):
                self._convert(directory, {'rootTag': '123 invalid', 'itemTag': 'item'})
            with self.assertRaisesRegex(ValueError, 'reserved XML prefix'):
                self._convert(directory, {'rootTag': 'root', 'itemTag': 'xmlItem'})

    def test_nested_json_keys_are_safely_normalized_to_valid_xml(self):
        with tempfile.TemporaryDirectory() as directory:
            output = self._convert(directory, {}, '{"-bad key": {"xmlThing": 1}}')
            root = ET.parse(output).getroot()
            self.assertEqual(root[0].tag, '_-bad_key')
            self.assertEqual(root[0][0].tag, '_xmlThing')


class SplitPdfChunkTests(unittest.TestCase):
    def _source_pdf(self, path: Path, page_count=10):
        writer = PdfWriter()
        for index in range(page_count):
            writer.add_blank_page(width=100 + index, height=200)
        with path.open('wb') as handle:
            writer.write(handle)

    def _zip_chunks(self, output: Path):
        chunks = []
        with zipfile.ZipFile(output) as archive:
            for name in archive.namelist():
                with archive.open(name) as handle:
                    reader = PdfReader(handle)
                    chunks.append((name, [int(page.mediabox.width) for page in reader.pages]))
        return chunks

    def test_ten_pages_split_into_ordered_groups_of_three(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / 'source.pdf'
            output = Path(directory) / 'chunks.zip'
            self._source_pdf(source)

            result = PdfCoreEngine.split([str(source)], str(output), {'mode': 'every_n', 'everyN': 3})

            self.assertEqual(result, str(output))
            chunks = self._zip_chunks(output)
            self.assertEqual([name for name, _ in chunks], ['pages_1-3.pdf', 'pages_4-6.pdf', 'pages_7-9.pdf', 'page_10.pdf'])
            self.assertEqual([len(widths) for _, widths in chunks], [3, 3, 3, 1])
            self.assertEqual([width for _, widths in chunks for width in widths], list(range(100, 110)))

    def test_one_page_chunks_and_n_larger_than_document(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / 'source.pdf'
            self._source_pdf(source, page_count=4)

            one_output = Path(directory) / 'one.zip'
            PdfCoreEngine.split([str(source)], str(one_output), {'mode': 'every_n', 'everyN': 1})
            self.assertEqual([len(pages) for _, pages in self._zip_chunks(one_output)], [1, 1, 1, 1])

            large_output = Path(directory) / 'large.zip'
            result = PdfCoreEngine.split([str(source)], str(large_output), {'mode': 'every_n', 'everyN': 20})
            self.assertTrue(result.endswith('pages_1-4.pdf'))
            self.assertEqual(len(PdfReader(result).pages), 4)

    def test_invalid_n_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / 'source.pdf'
            self._source_pdf(source, page_count=2)
            for value in (0, -1, 'not-a-number'):
                with self.subTest(value=value):
                    with self.assertRaisesRegex(Exception, 'positive integer|greater than zero'):
                        PdfCoreEngine.split([str(source)], str(Path(directory) / 'out.zip'), {'mode': 'every_n', 'everyN': value})


if __name__ == '__main__':
    unittest.main()
